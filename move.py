from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('bambo.xlsx')
ws = wb.active

ws.move_range("C1:D8" , rows=2 , cols=2)



wb.save('bambo.xlsx')