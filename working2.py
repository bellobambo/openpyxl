from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('bambo.xlsx')
ws = wb.active

# looping through rows
for row in range(1, 11):
    for col in range(1,7):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)


wb.save('bambo.xlsx')